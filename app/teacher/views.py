import shutil
import os
import zipfile
import json
import uuid
from flask import flash, redirect, render_template, url_for, request,\
    current_app, send_from_directory, make_response, send_file
from datetime import datetime
from flask_login import login_required, current_user
from ..upload_utils import secure_filename
from . import teacher
from .. import db
from ..auths import UserAuth
from .forms import up_corrected, UploadCorrected,\
    CourseForm, HomeworkForm, UploadResourceForm, upsr, AcceptTeam, RejectTeam, MoveForm
from ..models.models import Course, Homework, Team,\
    TeamMember, Student, Submission, Attachment, Teacher
from flask_uploads import UploadNotAllowed
from openpyxl.utils.exceptions import InvalidFileException
from config import basedir
from openpyxl import Workbook
from sqlalchemy import not_


@teacher.before_request
@login_required
def before_request():
    pass


@teacher.route('/<course_id>/course', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def set_course_info(course_id):
    # 教师课程信息
    form = CourseForm()
    course = Course.query.filter_by(id=course_id).first()
    if form.validate_on_submit():
        course.outline = form.outline.data
        course.teamsize_min = form.teamsize_min.data
        course.teamsize_max = form.teamsize_max.data
        db.session.add(course)
        db.session.commit()
        flash('修改成功！', 'success')
        return redirect(url_for('teacher.set_course_info', course_id=course_id))
    form.outline.data = course.outline
    form.teamsize_min.data = course.teamsize_min
    form.teamsize_max.data = course.teamsize_max
    return render_template('teacher/course.html', course_id=course_id, form=form, course=course)


@teacher.route('/<course_id>/resource', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def manage_resource(course_id):
    # 教师查看修改课程资源
    path = request.args.get('path')
    if not path:
        return redirect(url_for('teacher.manage_resource', course_id=course_id, path='/'))

    expand_path = os.path.join(current_app.config['UPLOADED_FILES_DEST'], 'resource', course_id, path[1:])
    if not os.path.exists(expand_path):
        # 没有文件夹？赶紧新建一个，真鸡儿丢人
        os.mkdir(expand_path)

    if request.method == 'POST' and 'file' in request.files:
        # 上传文件
        file = request.files['file']
        filename = secure_filename(file.filename)
        filepath = os.path.join(current_app.config['UPLOADED_FILES_DEST'], 'resource', course_id, path[1:], filename)
        if os.path.exists(filepath):
            flash('已经存在了同名文件', 'danger')
            return redirect(url_for('teacher.manage_resource', course_id=course_id, path=path))
        file.save(filepath)
        flash('上传成功！', 'success')
        return redirect(url_for('teacher.manage_resource', course_id=course_id, path=path))

    if request.method == 'POST' and 'dirname' in request.form:
        # 新建文件夹
        dirname = request.form['dirname']
        dirpath = os.path.join(current_app.config['UPLOADED_FILES_DEST'], 'resource', course_id, path[1:], dirname)
        if os.path.exists(dirpath):
            flash('已经存在了同名文件夹', 'danger')
            return redirect(url_for('teacher.manage_resource', course_id=course_id, path=path))
        os.mkdir(dirpath)
        flash('新建文件夹成功！', 'success')
        return redirect(url_for('teacher.manage_resource', course_id=course_id, path=path))

    if 'action' in request.form:
        if request.form.get('action') == 'delete':
            # 删除
            filepath = os.path.join(
                current_app.config['UPLOADED_FILES_DEST'],
                'resource',
                course_id,
                path[1:],
                request.form.get('filename')
            )
            if not os.path.exists(filepath):
                flash('文件不存在！', 'danger')
                return redirect(url_for('teacher.manage_resource', course_id=course_id, path=path))
            if os.path.isfile(filepath):
                os.remove(filepath)
            else:
                shutil.rmtree(filepath)
            flash('删除成功！', 'success')
            return redirect(url_for('teacher.manage_resource', course_id=course_id, path=path))

    if request.args.get('download'):
        # 下载
        filedir = os.path.join(
            current_app.config['UPLOADED_FILES_DEST'],
            'resource',
            course_id,
            path[1:])
        filename = request.args.get('filename')
        print(filename)
        if os.path.exists(os.path.join(filedir, filename)):
            return send_from_directory(filedir, filename, as_attachment=True)
        else:
            flash('文件不存在！', 'danger')
            return redirect(url_for('teacher.manage_resource', course_id=course_id, path=path))

    files = []

    def sizeof_fmt(num, suffix='B'):
        for unit in ['', 'K', 'M', 'G', 'T', 'P', 'E', 'Z']:
            if abs(num) < 1024.0:
                return "%3.1f%s%s" % (num, unit, suffix)
            num /= 1024.0
        return "%.1f%s%s" % (num, 'Y', suffix)

    class file_attributes:
        name = ""
        size = ""
        create_time = datetime.min
        is_dir = False
        is_file = False

        def __init__(self, name, size, create_time, is_dir, is_file):
            self.name = name
            self.size = size
            self.create_time = create_time
            self.is_dir = is_dir
            self.is_file = is_file

    for file in os.scandir(expand_path):
        time = datetime.fromtimestamp(file.stat().st_mtime)
        files.append(file_attributes(file.name, sizeof_fmt(file.stat().st_size), time, file.is_dir(), file.is_file()))
    return render_template('teacher/resource.html', course_id=course_id, files=files, path=path)


@teacher.route('/<course_id>/homework', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def homework(course_id):

    form = HomeworkForm()

    if request.args.get('get_teamhomework_all'):
        return get_teamhomework_all(course_id)
    elif request.args.get('get_score_all'):
        return get_score_all(course_id)

    if form.validate_on_submit():
        if form.weight.data > 100 or form.weight.data <= 0:
            flash('无效的权重', 'danger')
            return redirect(url_for('teacher.homework', course_id=course_id))
        begin_time, end_time = form.time.data.split(' - ')
        begin_time = datetime.strptime(begin_time, '%m/%d/%Y %H:%M')
        end_time = datetime.strptime(end_time, '%m/%d/%Y %H:%M')
        homework = Homework()

        homework.name = form.name.data
        homework.base_requirement = form.base_requirement.data
        homework.begin_time = begin_time
        homework.end_time = end_time
        homework.weight = form.weight.data
        homework.max_submit_attempts = form.max_submit_attempts.data
        homework.course_id = course_id

        db.session.add(homework)
        db.session.commit()

        flash('发布成功！', 'success')
        return redirect(url_for('teacher.homework', course_id=course_id))
    course = Course.query.filter_by(id=course_id).first()

    homework_list = Homework.query.filter_by(course_id=course_id).all()
    return render_template('teacher/homework.html', course_id=course_id, homeworks=homework_list, form=form, course=course)


# PudgeG负责：提交情况表导出↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓
def get_teamhomework_all(course_id):
    # 得到所有小队历次作业提交信息
    workbook = Workbook()
    worksheet = workbook.create_chartsheet()
    worksheet.title = '团队累计提交作业情况'

    team_list = Team.query.filter_by(course_id=course_id).filter_by(status=2).all()
    Team.team_list(course_id)
    homework_list = Homework.query.filter_by(course_id=course_id).all()
    Homework.homework_list(course_id)

    # 总列数
    column_number = 2
    # 总行数
    row_number = 1

    # 第一行输入的内容
    for every_homework in homework_list:
        column_number += 1
        worksheet.cell(row=1, column=column_number).value = '第' + every_homework.order + '次作业成绩'
        worksheet.cell(row=1, column=1).value = '团队名称'
        worksheet.cell(row=1, column=2).value = '团队编号'

    # 后续内容循环输入
    for every_team in team_list:
        row_number += 1
        worksheet.cell(row=row_number, column=1).value = every_team.team_name
        worksheet.cell(row=row_number, column=2).value = every_team.order
        i = 2
        for every_homework in homework_list:
            i += 1
            _submission = Submission.query.filter_by(homework_id=every_homework.id).filter_by(team_id=every_team.id).all()
            if _submission is None:
                worksheet.cell(row=row_number, column=i).value = '0'
            else:
                submission = Submission.query.filter_by(homework_id=every_homework.id).filter_by(team_id=every_team.id).all()[-1]
                worksheet.cell(row=row_number, column=i).value = submission.score

    workbook.save('team_homework_all.xlsx')
    if os.path.isfile(os.path.join(os.getcwd(), 'homework', 'team_homework_all.xlsx')):
        response = make_response(send_file(os.path.join(os.getcwd(), 'homework', 'team_homework_all.xlsx')))
    else:
        flash('文件创建失败！', 'danger')
        return redirect(url_for('teacher.homework', course_id=course_id))
    response.headers["Content-Disposition"] = "attachment; filename=" + 'team_homework_all.xlsx' + ";"
    return response
# PudgeG负责：提交情况表导出↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑


# PudgeG负责：总成绩表导出↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓
def get_score_all(course_id):
    # 得到小队总成绩以及个人总成绩
    workbook = Workbook()

    worksheet_team = workbook.create_chartsheet()
    worksheet_team.title = '小队总成绩'
    worksheet_team.append('团队名称', '团队编号', '团队总成绩')

    worksheet_member = workbook.create_chartsheet()
    worksheet_member.title = '成员成绩'
    worksheet_member.append('团队名称', '团队编号', '学生姓名', '学生编号', '学生总成绩')

    team_list = Team.query.filter_by(course_id=course_id).filter_by(status=2).all()
    Team.team_list(course_id)
    homework_list = Homework.query.filter_by(course_id=course_id).all()
    Homework.homework_list(course_id)
    input_info = []
    for team in team_list:
        # Team表
        score = 0
        for every_homework in homework_list:
            _this_submission = Submission.query.filter_by(team_id=team.id).filter_by(homework_id=every_homework.id).first()
            if _this_submission is not None:
                this_submission = Submission.query.filter_by(team_id=team.id).filter_by(homework_id=every_homework.id).all()[-1]
                score += this_submission.score * every_homework.weight / 100
        submission_record = {'团队名称': team.team_name,
                             '团队编号': team.order,
                             '团队总成绩': score}
        input_info.append(submission_record)
    worksheet_team.append(input_info)

    input_info2 = []
    for team in team_list:
        # Member表
        member_list = TeamMember.query.filter_by(team_id=team.id).filter_by(status=1).all()

        # 团队负责人第一行
        owner = Student.query.filter_by(id=team.owner_id).first()
        score = 0
        for every_homework in homework_list:
            _this_submission = Submission.query.filter_by(team_id=team.id).filter_by(homework_id=every_homework.id).first()
            if _this_submission is not None:
                this_submission = Submission.query.filter_by(team_id=team.id).filter_by(homework_id=every_homework.id).all()[-1]
                score += this_submission.score * every_homework.weight / 100
        submission_record = {'团队名称': team.team_name,
                             '团队编号': team.order,
                             '学生姓名': owner.name,
                             '学生编号': owner.id,
                             '学生总成绩': score * team.owner_grade}
        input_info2.append(submission_record)

        for every_member in member_list:
            _every_member = Student.query.filter_by(id=every_member.student_id).first()
            submission_record = {'团队名称': team.team_name,
                                 '团队编号': team.order,
                                 '学生姓名': _every_member.name,
                                 '学生编号': _every_member.id,
                                 '学生总成绩': score * every_member.grade}
            input_info2.append(submission_record)
    worksheet_member.append(input_info2)

    workbook.save('score_all.xlsx')
    if os.path.isfile(os.path.join(os.getcwd(), 'homework', 'score_all.xlsx')):
        response = make_response(send_file(os.path.join(os.getcwd(), 'homework', 'score_all.xlsx')))
    else:
        flash('文件创建失败！', 'danger')
        return redirect(url_for('teacher.homework', course_id=course_id))
    response.headers["Content-Disposition"] = "attachment; filename=" + 'score_all.xlsx' + ";"
    return response
# PudgeG负责：总成绩表导出↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑


@teacher.route('/<course_id>/homework/<homework_id>', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def homework_detail(course_id, homework_id):
    # 教师查看作业详情&团队提交状况
    form = HomeworkForm()
    course = Course.query.filter_by(id=course_id).first()
    homework = Homework.query.filter_by(id=homework_id).first()

    if request.args.get('homework_report'):
        return get_homework_report(homework_id)

    if form.validate_on_submit():
        # 修改作业
        begin_time, end_time = form.time.data.split(' - ')
        begin_time = datetime.strptime(begin_time, '%m/%d/%Y %H:%M')
        end_time = datetime.strptime(end_time, '%m/%d/%Y %H:%M')

        homework.name = form.name.data
        homework.base_requirement = form.base_requirement.data
        homework.begin_time = begin_time
        homework.end_time = end_time
        homework.weight = form.weight.data
        homework.max_submit_attempts = form.max_submit_attempts.data
        homework.course_id = course_id

        flash('修改成功！', 'success')
        return redirect(url_for('teacher.homework_detail',
                                course_id=course_id,
                                homework_id=homework_id))

    form.name.data = homework.name
    form.base_requirement.data = homework.base_requirement
    form.time.data = '{} - {}'.format(homework.begin_time.strftime('%m/%d/%Y %H:%M'),
                                      homework.end_time.strftime('%m/%d/%Y %H:%M'))
    form.weight.data = homework.weight
    form.max_submit_attempts.data = homework.max_submit_attempts

    return render_template('teacher/homework_detail.html',
                           course_id=course_id,
                           course=course,
                           form=form,
                           homework=homework)


# PudgeG负责：得到本次作业报表↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓
def get_homework_report(homework_id):
    # 得到本次作业报表
    this_homework = Homework.query.filter_by(id=homework_id).first()
    team_this_course = Team.query.filter_by(course_id=this_homework.course_id).filter_by(status=2).all()
    Team.team_list(this_homework.course_id)

    # submission_list = Submission.query.filter_by(homework_id=homework_id).all()
    # if len(submission_list) == 0:
    #     flash('无提交记录，请先催交！', 'danger')
    #     return redirect(request.args.get('next') or url_for('main.set_homework'))

    workbook = Workbook()
    worksheet = workbook.create_chartsheet()
    worksheet.title = this_homework.name + ' 提交情况'
    worksheet.append(['团队名称', '团队ID', '本次作业是否提交', '本次作业分数'])
    input_info = []
    for team in team_this_course:
        _finished = Submission.query.filter_by(homework_id=homework_id).filter_by(team_id=team.id).all()
        if _finished is None:
            # 无提交记录
            homework_record = {'团队名称': team.team_name,
                               '团队ID': team.order,
                               '本次作业是否提交': 'No',
                               '本次作业分数': 0}
        else:
            # 有提交记录，拿最后一个
            finished = Submission.query.filter_by(homework_id=homework_id).filter_by(team_id=team.id).all()[-1]
            homework_record = {'团队名称': team.team_name,
                               '团队ID': team.order,
                               '本次作业是否提交': 'Yes',
                               '本次作业分数': finished.score}
        input_info.append(homework_record)

        # def convert_status(status):
        #     switcher = {
        #         0: '作业未批改',
        #         1: '作业已批改'
        #     }
        #     return switcher.get(status, '其他')

    worksheet.append(input_info)
    workbook.save('homework_report.xlsx')
    if os.path.isfile(os.path.join(os.getcwd(), 'homework', 'homework_report.xlsx')):
        response = make_response(send_file(os.path.join(os.getcwd(), 'homework', 'homework_report.xlsx')))
    else:
        flash('文件创建失败！', 'danger')
        return redirect(url_for('teacher/teacher_teammanagement'))
    response.headers["Content-Disposition"] = "attachment; filename=" + 'homework_report.xlsx' + ";"
    return response
# PudgeG负责：得到本次作业报表↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑


@teacher.route('/uploadresource', methods=['GET', 'POST'])
def teacher_resource():
    form = UploadResourceForm()
    if form.validate_on_submit():
        try:
            (name, ext) = os.path.splitext(form.up.data.filename)
            filename = upsr.save(form.up.data, basedir + '/uploads/teacher_resources', name=str(uuid.uuid4())+ext)
            file_url = upsr.url(filename)
        except UploadNotAllowed:
            flash('附件上传不允许！', 'danger')
            return redirect(request.args.get('next') or url_for('uploadresource.html'))
    else:
        file_url = None
    return render_template('uploadresource.html', form=form, file_url=file_url)


# 上传老师批改后的作业
def teacher_corrected(course_id, homework_id):
    form = UploadCorrected()
    if form.validate_on_submit():
        if form.up_corrected:
            try:
                name_temp, ext = os.path.splitext(form.up_corrected.data.filename)
                # 保存文件在basedir/uploads/<course_id>/<homework_id>/teacher_corrected.ext (通过特定名字找下载)
                up_corrected.save(form.up_corrected,
                                  folder=os.path.join(basedir, 'uploads', str(course_id), str(homework_id)),
                                  name=u'teacher_corrected' + ext)
            except InvalidFileException:
                flash(u'附件类型不正确，请使用zip或rar', 'danger')
                return redirect(request.args.get('next') or url_for('teacher.teacher_corrected'))
            except UploadNotAllowed:
                flash(u'附件上传不允许')
                return redirect(request.args.get('next') or url_for('teacher.teacher_corrected'))
            # 可能加入全体广播 向全部学生广播教师修改作业已上传
            flash('上传成功')
            return redirect(url_for('teacher.teacher_corrected', form=form))
    return render_template('teacher/upload_corrected.html', form=form)


def add_member(student_id, team_id):
    team_member = TeamMember()
    team_member.team_id = team_id
    team_member.student_id = student_id
    team_member.status = 1
    db.session.add(team_member)
    delete_list = TeamMember.query.filter_by(status=2).filter_by(student_id=student_id).all()
    for a in delete_list:
        db.session.delete(a)
    db.session.commit()

# PudgeG负责:团队报表导出↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓
def get_team_report(course_id):
    down_list = Team.query.filter_by(course_id=course_id).filter_by(status=2).all()
    Team.team_list(course_id)
    if down_list is None:
        flash('没有已接受团队，请等待申请并批准！', 'danger')
        return redirect(request.args.get('next') or url_for('main.teacher_teammanagement'))
    workbook = Workbook()
    worksheet = workbook.create_chartsheet()
    worksheet.title = '已接受团队信息'
    worksheet.append(['队伍名称', '队伍编号', '成员姓名', '成员编号', '成员角色'])
    # i = 0 表示队伍数量
    input_info = []
    for team in down_list:
        member_list = TeamMember.query.filter_by(team_id=team.id).all()
        input_record = {'队伍名称': team.team_name,
                        '队伍编号': team.order,
                        '成员姓名': Student.query.filter_by(id=team.owner_id).name,
                        '成员编号': team.owner_id,
                        '成员角色': '团队负责人'}
        input_info.append(input_record)
        # num_of_member = len(member_list)+1 表示每支队伍人员数量
        # i += 1
        for member in member_list:
            input_record = {'队伍名称': team.team_name,
                            '队伍编号': team.order,
                            '成员姓名': Student.query.filter_by(id=member.student_id).name,
                            '成员编号': member.student_id,
                            '成员角色': '普通成员'}
            input_info.append(input_record)
    worksheet.append(input_info)
    workbook.save('team_table.xlsx')
    if os.path.isfile(os.path.join(os.getcwd(), 'team_manage', 'team_table.xlsx')):
        response = make_response(send_file(os.path.join(os.getcwd(), 'team_manage', 'team_table.xlsx')))
    else:
        flash('文件创建失败！', 'danger')
        return redirect(url_for('teacher/teacher_teammanagement'))
    response.headers["Content-Disposition"] = "attachment; filename=" + 'team_table.xlsx' + ";"
    return response
# PudgeG负责:团队报表输出↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑↑


# 用于生成zip文件
def make_zip(source_dir, output_filename):
    zipf = zipfile.ZipFile(output_filename, 'w')
    pre_len = len(os.path.dirname(source_dir))
    for parent, dirnames, filenames in os.walk(source_dir):
        for filename in filenames:
            pathfile = os.path.join(parent, filename)
            # 相对路径
            arcname = pathfile[pre_len:].strip(os.path.sep)
            zipf.write(pathfile, arcname)
    zipf.close()


# 用于下载前对文件重命名
def rename(source_dir, rename_dic):
    for parent, dirnames, filenames in os.walk(source_dir):
        for filename in filenames:
            name, ext = os.path.splitext(filename)
            if name in rename_dic.keys():
                os.rename(os.path.join(parent, filename), os.path.join(parent, rename_dic[name]))


@teacher.route('/teacher/<course_id>/givegrade_teacher/<homework_id>', methods=['GET', 'POST'])
def givegrade_teacher(course_id, homework_id):

    # 显示学生已提交的作业(显示最新的提交记录)
    submission = Submission.query.filter_by(homework_id=homework_id).filter_by(submit_status=1).all()
    team = Team.query.filter_by(course_id=course_id).all()

    #取每个组最新的提交记录
    submission_list = []
    for team_temp in team:
        submission_latest = Submission \
            .query \
            .filter_by(team_id=team_temp.id,
                       homework_id=homework_id) \
            .order_by(Submission.id.desc()) \
            .first()
        submission_list.append(submission_latest)

    homework_list = []
    for i in submission_list:
        team = Team.query.filter_by(id=i.team_id).first()
        homework_list.append({'team_id': i.team_id, 'team_name': team.team_name, 'text_content': i.text_content,
                              'score': i.score, 'comments': i.comments})

    # json [{'team_id':team_id, 'score': score, 'comments': comments}]
    # 提交评价和评论
    if request.method == 'POST' and request.form.get('action') == 'submit':
        _list = json.loads(request.form.get('data'))
        for dic in _list:
            for submission_temp in submission_list:
                if submission_temp.team_id == dic['team_id']:
                    submission_temp.score = dic['score']
                    submission_temp.comments = dic['comments']
                    db.session.add(submission_temp)
        db.session.commit()
        return redirect(url_for('teacher.givegrade_teacher', course_id=course_id, homework_id=homework_id))

    # 单个下载学生作业
    if request.method == 'POST' and request.form.get('action') == 'download':

        team_id = request.form.get('team_id')
        file_dir = os.path.join(current_app.config['UPLOADED_FILES_DEST'],
                                str(course_id),
                                str(homework_id),
                                str(team_id))

        # 取最新的一次上传和上传时的附件
        submission_previous = Submission \
            .query \
            .filter_by(team_id=team_id,
                       homework_id=homework_id) \
            .order_by(Submission.id.desc()) \
            .first()

        attachment_previous = None
        if submission_previous:
            attachment_previous = Attachment.query.filter_by(submission_id=submission_previous.id).first()

        # 无附件
        if not attachment_previous:
            flash('该组没有上传作业附件')
            return redirect(url_for('teacher.givegrade_teacher', course_id=course_id, homework_id=homework_id))
        else:

            filename_upload = attachment_previous.file_name
            file_uuid = attachment_previous.guid

            # 寻找保存目录下的uuid文件
            for i in os.listdir(file_dir):
                if i.startswith(str(file_uuid)):
                    os.rename(i, filename_upload)
            return send_from_directory(directory=file_dir, filename=filename_upload, as_attachment=True)

    # 批量下载学生作业
    if request.method == 'POST' and request.form.get('action') == 'multi_download':

        file_path = os.path.join(basedir, 'uploads', str(course_id), str(homework_id))
        save_path = os.path.join(basedir, 'temp', 'download.zip')

        submission_all = Submission \
            .query \
            .filter_by(homework_id=homework_id) \
            .order_by(Submission.id.desc()) \
            .all()

        # 建立 uuid与 上传时file_name的 键值对关系{uuid: file_name}
        rename_list = {}
        for submission_temp in submission_all:
            attachment_temp = Attachment.query.filter_by(submission_id=submission_temp.id).first()
            rename_list[str(attachment_temp.guid)] = str(attachment_temp.file_name)

        # 重命名文件并提供下载
        rename(file_path, rename_list)
        make_zip(file_path, save_path)
        return send_from_directory(directory=file_path, filename='download.zip', as_attachment=True)
    return render_template('teacher/givegrade_teacher.html', homework_list=homework_list)


# 教师查看往期课程和往期课程作业
@teacher.route('/see_class_before', methods=['GET', 'POST'])
def see_class_before():
    #取出当前老师参加的所有课程
    course = Teacher.query.filter_by(id=current_user.id).course

    course_detail_info_list = []
    # 显示往期课程信息
    for i in course:
        course_detail_info_list.append({'course_id': i.id, 'course_name': i.name, 'course_credit': i.credit,
                                        'course_student_number': len(i.student), 'course_info': i.course_info})

    # 下载往期课程作业
    if request.method == 'POST' and request.form.get('action') == 'download':

        course_id = request.args.get('course_id')
        file_path = os.path.join(basedir, 'uploads', str(course_id))
        save_path = os.path.join(basedir, 'temp', 'download.zip')

        if os.path.exists(file_path):
            make_zip(file_path, save_path)
            return send_from_directory(directory=file_path, filename='download.zip', as_attachement=True)
        else:
            flash('这个课程没有附件作业保存！', 'danger')
            return redirect(url_for('teacher.see_class_before'))
    return render_template('teacher/see_class_before.html', course_detail_info_list=course_detail_info_list)


@teacher.route('/<course_id>/team', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def team_manage(course_id):
    # 教师管理团队
    course = Course.query.filter_by(id=course_id).first()
    teams = Team.query.filter_by(course_id=course_id).all()
    form = MoveForm()
    if request.form.get('action') == 'accept':
        team = Team.query.filter_by(id=request.form.get('team_id')).first()
        if team:
            team.status = 2
            db.session.add(team)
            db.session.commit()
            flash("通过成功", "success")
            return redirect(url_for('teacher.team_manage', course_id=course_id))
        else:
            flash("找不到此团队", "danger")
            return redirect(url_for('teacher.team_manage', course_id=course_id))
    elif request.form.get('action') == 'reject':
        team = Team.query.filter_by(id=request.form.get('team_id')).first()
        if team:
            team.status = 3
            team.reject_reason = request.form.get('reason')
            db.session.add(team)
            db.session.commit()
            flash("拒绝成功", "success")
            return redirect(url_for('teacher.team_manage', course_id=course_id))
        else:
            flash("找不到此团队", "danger")
            return redirect(url_for('teacher.team_manage', course_id=course_id))

    unteamed_group = list(course.students)
    for team in teams:
        unteamed_group.remove(team.owner)

    members = TeamMember.query.filter(not_(TeamMember.status == 2)).filter(TeamMember.team.has(course_id=course_id)).all()

    for member in members:
        unteamed_group.remove(member.student)

    pending_teams = Team\
        .query\
        .filter_by(course_id=course_id)\
        .filter(not_(Team.status == 2))\
        .all()

    form.pending_teams.choices = [(r.id, r.team_name) for r in pending_teams]

    if form.validate_on_submit():
        print(2333)
        add_member(form.student.data, form.pending_teams.data)
        flash('移动成功！', 'success')
        return redirect(url_for('teacher.team_manage', course_id=course_id))

    return render_template('teacher/team.html',
                           course_id=course_id,
                           course=course,
                           teams=teams,
                           unteamed_group=unteamed_group,
                           form=form)
