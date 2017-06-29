from flask import render_template, redirect, url_for, flash, request, session
from . import main
from .forms import AddSemesterForm
from .. import db
from ..models.models import Semester
import os
from datetime import date
from .forms import AddSemesterForm, CourseForm, CourseFormTeacher, upsr, UploadResourceForm
from ..models.models import Student, Teacher, SCRelationship, TCRelationship, Course, Semester
from flask_login import current_user, login_required
from functools import wraps
from flask import request
from .. import config, ups
import openpyxl
from config import basedir

ALLOWED_EXTENSIONS = {"xls", "xlsx", "csv"}             # set(["xls", "xlsx"]) 允许上传的文件类型


class UserAuth:
    """
    权限认证装饰器，对应装饰器应用在route下即可，若权限不够将返回主页
    """
    @staticmethod
    def dean(func):
        @wraps(func)
        def decorated(*args, **kwargs):
            if not current_user.user_type() == 0:
                flash('无权限', 'danger')
                return redirect(url_for('main.index'))
            else:
                return func(*args, **kwargs)
        return decorated

    @staticmethod
    def teacher(func):
        @wraps(func)
        def decorated(*args, **kwargs):
            if not current_user.user_type() == 1:
                flash('无权限', 'danger')
                return redirect(url_for('main.index'))
            else:
                return func(*args, **kwargs)
        return decorated

    @staticmethod
    def student(func):
        @wraps(func)
        def decorated(*args, **kwargs):
            if not current_user.user_type() == 2:
                flash('无权限！', 'danger')
                return redirect(url_for('main.index'))
            else:
                return func(*args, **kwargs)
        return decorated

    @staticmethod
    def teacher_course_access(func):
        @wraps(func)
        def decorated(*args, **kwargs):
            if not TCRelationship.query.filter_by(teacher_id=current_user.id, course_id=kwargs['course_id']).first():
                flash('无权限！', 'danger')
                return redirect(url_for('main.index'))
            else:
                return func(*args, **kwargs)
        return decorated

    @staticmethod
    def student_course_access(func):
        @wraps(func)
        def decorated(*args, **kwargs):
            if not SCRelationship.query.filter_by(student_id=current_user.id, course_id=kwargs['course_id']).first():
                flash('无权限！', 'danger')
                return redirect(url_for('main.index'))
            else:
                return func(*args, **kwargs)
        return decorated


@main.before_request
@login_required
def before_request():
    pass


@main.route('/', methods=['GET', 'POST'])
def index():

    #教师特色主页
    if current_user.user_type() == 1:
        tcrels = TCRelationship.query.filter_by(teacher_id=current_user.id).all()
        courses = []
        for rel in tcrels:
            courses.append(Course.query.filter_by(id=rel.course_id).first())
        return render_template('teacher/index.html', courses=courses)

    # 学生特色主页
    if current_user.user_type() == 2:
        tcrels = SCRelationship.query.filter_by(student_id=current_user.id).all()
        courses = []
        for rel in tcrels:
            courses.append(Course.query.filter_by(id=rel.course_id).first())
        return render_template('student/index.html', courses=courses)

    return render_template('index.html')


@main.route('/manage/semester', methods=['GET', 'POST'])
@UserAuth.dean
def manage_semester():
    form = AddSemesterForm()
    if form.validate_on_submit():
        begin_time, end_time = form.time.data.split('-')
        month, day, year = begin_time.split('/')
        begin_time = date(int(year), int(month), int(day))
        month, day, year = end_time.split('/')
        end_time = date(int(year), int(month), int(day))
        if Semester.query.filter_by(id=form.id.data).first():
            flash('添加了重复的学期…', 'danger')
            return redirect(url_for('main.manage_semester'))
        db.session.add(Semester(id=form.id.data, base_info=form.base_info.data,
                                begin_time=begin_time, end_time=end_time))
        db.session.commit()
        flash('添加成功！', 'success')
        return redirect(url_for('main.manage_semester'))
    semester_list = Semester.query.all()
    return render_template('manage/semester.html', form=form, semesters=semester_list)


def read_file(file_path):
    workbook = openpyxl.load_workbook(filename=file_path)  # 打开xls文件

    sheet_student = workbook.get_sheet_by_name('学生信息')
    sheet_teacher = workbook.get_sheet_by_name('老师信息')  # 通过sheet名字访问sheet
    student_info = []
    teacher_info = []
    for i in range(2, sheet_student.max_row + 1):
        student_list = {'id': sheet_student.cell(row=i, column=1).value,
                        'name': sheet_student.cell(row=i, column=2).value,
                        'password': 666}  # 学生初始密码 666
        student_info.append(student_list)
    for i in range(2, sheet_teacher.max_row + 1):
        teacher_list = {'id': sheet_teacher.cell(row=i, column=1).value,
                        'name': sheet_teacher.cell(row=i, column=2).value,
                        'teacher_info': sheet_teacher.cell(row=i, column=3).value,
                        'password': 666}  # 老师初始密码 666
        teacher_info.append(teacher_list)
    return student_info, teacher_info


@main.route('/index-teacher', methods=['GET', 'POST'])
def index_teacher():
    return render_template('auth_teacher/index_teacher.html')


@main.route('/index-teacher/teacher-course', methods=['GET', 'POST'])
def teacher_course():
    return render_template('auth_teacher/teacher_course.html')


@main.route('/uploadresource', methods=['GET', 'POST'])
def teacher_resource():  # TODO: add 文件系统
    form = UploadResourceForm()
    if form.validate_on_submit():
        filename = upsr.save(form.up.data)
        file_url = upsr.url(filename)
    else:
        file_url = None
    return render_template('uploadresource.html', form=form, file_url=file_url)


@main.route('/index-teacher/teacher-homework', methods=['GET', 'POST'])
def teacher_homework():
    return render_template('auth_teacher/teacher_homework.html')


@main.route('/index-teacher/teacher-communicate', methods=['GET', 'POST'])
def teacher_communicate():
    return render_template('auth_teacher/teacher_communicate.html')


@main.route('/index-teacher/teacher-teammanagement', methods=['GET', 'POST'])
def teacher_teammanagement():
    return render_template('auth_teacher/teacher_teammanagement.html')


@main.route('/manage/course', methods=['GET', 'POST'])
@UserAuth.dean
def manage_course():
    semester_list = Semester.query.all()
    form = CourseForm()
    form.semester.choices = [(a.id, str(a.id / 100) + '学年第' + str(a.id % 100) + '学期') for a in semester_list]
    if 'action' in request.args:
        if request.args['action'] == 'delete':
            _course = Course.query.filter_by(id=int(request.args['id'])).first()
            if not _course:
                flash('找不到该课程', 'danger')
                return redirect(request.args.get('next') or url_for('main.manage_course'))
            db.session.delete(_course)
            db.session.commit()
            flash('删除成功', 'success')
            return redirect(request.args.get('next') or url_for('main.manage_course'))
        elif request.args['action'] == 'end':
            _course = Course.query.filter_by(id=int(request.args['id'])).first()
            if not _course:
                flash('找不到该课程', 'danger')
                return redirect(request.args.get('next') or url_for('main.manage_course'))
            _course.status = False
            db.session.commit()
            flash('结束成功', 'success')
            return redirect(request.args.get('next') or url_for('main.manage_course'))

    if form.validate_on_submit():

        # course的基本信息
        course = Course()
        course.name = form.name.data
        course.course_info = form.course_info.data
        course.place = form.place.data
        course.credit = int(form.credit.data)
        course.semester_id = form.semester.data
        course.outline = '无'
        course.teamsize_min = 1
        course.teamsize_max = 5
        course.status = True

        db.session.add(course)
        db.session.commit()

        # 上传文件处理
        filename = ups.save(form.stuff_info.data)
        file_path = os.path.join(basedir, 'uploads', filename)
        student_info, teacher_info = read_file(file_path=file_path)

        # 添加学生
        for i in student_info:
            student = Student.query.filter_by(id=i.get('id')).first()
            if student is None:
                student = Student()
            student.id = int(i.get('id'))
            student.name = i.get('name')
            student.password = str(i.get('password'))
            db.session.add(student)

            # 添加学生课程关系
            screl = SCRelationship.query.filter_by(student_id=student.id, course_id=course.id).first()
            if screl is None:
                screl = SCRelationship(student_id=student.id, course_id=course.id)
            db.session.add(screl)

        # 添加教师
        for i in teacher_info:
            teacher = Teacher.query.filter_by(id=i.get('id')).first()
            if teacher is None:
                teacher = Teacher()
            teacher.id = int(i.get('id'))
            teacher.name = i.get('name')
            teacher.teacher_info = i.get('teacher_info')
            teacher.password = str(i.get('password'))
            db.session.add(teacher)

            # 添加老师课程关系
            tcrel= TCRelationship.query.filter_by(teacher_id=teacher.id, course_id=course.id).first()
            if tcrel is None:
                tcrel = TCRelationship(teacher_id=teacher.id, course_id=course.id)
            db.session.add(tcrel)

        db.session.commit()
        os.remove(file_path)

        flash('添加成功！', 'success')
        return redirect(request.args.get('next') or url_for('main.manage_course'))

    course_list = Course.query.all()  # 显示课程
    stuff_list = {}
    for course in course_list:
        sclist = SCRelationship.query.filter_by(course_id=course.id).all()
        student_list = []
        for screl in sclist:
            student_list.extend(Student.query.filter_by(id=screl.student_id).all())
        tclist = TCRelationship.query.filter_by(course_id=course.id).all()
        teacher_list = []
        for tcrel in tclist:
            teacher_list.extend(Teacher.query.filter_by(id=tcrel.teacher_id).all())
        stuff_list[course.id] = {
            'student_list': student_list,
            'teacher_list': teacher_list
        }

    return render_template('manage/course.html', form=form, courses=course_list, semesters=semester_list, stuff=stuff_list)


@main.route('/teacher/<course_id>/course', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def set_course_info(course_id):
    form = CourseFormTeacher()
    course = Course.query.filter_by(id=course_id).first()
    if form.validate_on_submit():
        course.outline = form.outline.data
        course.teamsize_min = form.teamsize_min.data
        course.teamsize_max = form.teamsize_max.data
        db.session.add(course)
        db.session.commit()
        flash('修改成功！', 'success')
        return redirect(url_for('main.set_course_info', course_id=course_id))
    form.outline.data = course.outline
    form.teamsize_min.data = course.teamsize_min
    form.teamsize_max.data = course.teamsize_max
    return render_template('teacher/course.html', course_id=course_id, form=form, course=course)


@main.route('/teacher/<course_id>/resource', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def manage_resource(course_id):
    return render_template('teacher/resource.html', course_id=course_id)


@main.route('/teacher/<course_id>/homework', methods=['GET', 'POST'])
@UserAuth.teacher_course_access
def set_homework(course_id):
    return render_template('teacher/homework.html', course_id=course_id)


@main.route('/student/<course_id>/course', methods=['GET'])
@UserAuth.student_course_access
def show_course(course_id):
    course = Course.query.filter_by(id=course_id).first()
    return render_template('student/course.html', course_id=course_id, course=course)

